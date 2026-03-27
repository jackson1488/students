"""
Модуль `app/routes/exports.py`

Назначение:
- API экспорта данных (оценки, расписание, ДЗ) в HTML/PDF/XLSX.

Ключевые константы и значения:
- `ALLOWED_FORMATS`: используется как конфигурационный или справочный набор значений в этом модуле.
- `ALLOWED_LANGUAGES`: используется как конфигурационный или справочный набор значений в этом модуле.
- `DAY_CANONICAL`: используется как конфигурационный или справочный набор значений в этом модуле.
- `DAY_ORDER`: используется как конфигурационный или справочный набор значений в этом модуле.
- `DAY_TRANSLATIONS`: используется как конфигурационный или справочный набор значений в этом модуле.
- `EXPORT_TEXT`: используется как конфигурационный или справочный набор значений в этом модуле.

Маршруты HTTP (если это route-модуль):
- `GET /exports/grades/<student_ref>` -> `export_grades`
- `GET /exports/schedule/<group_id>` -> `export_schedule`
- `GET /exports/homework/<int:homework_id>` -> `export_homework`

Функции модуля:
- `_get_format`: Возвращает данные по запросу.
- `_get_lang`: Возвращает данные по запросу.
- `_tr`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `_localize_day`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `_error_response`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `_extract_bearer_token`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `_resolve_current_user`: Определяет/находит нужный объект по входным параметрам.
- `_profile_name`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `_build_html`: Формирует служебную структуру данных/ответ.
- `_build_excel`: Формирует служебную структуру данных/ответ.
- `_resolve_pdf_font_name`: Определяет/находит нужный объект по входным параметрам.
- `_build_pdf`: Формирует служебную структуру данных/ответ.
- `_build_file_bytes`: Формирует служебную структуру данных/ответ.
- `_file_meta`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `_schedule_day_rank`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `export_grades`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `export_schedule`: Содержит бизнес-логику модуля и используется в общем потоке работы API.
- `export_homework`: Содержит бизнес-логику модуля и используется в общем потоке работы API.

Примечание:
- Комментарии в этом модуле ориентированы на чтение кода новичком: сначала цель, затем ключевые значения, потом функции.
"""

import html
import os
from datetime import datetime
from io import BytesIO

import jwt
from flask import Blueprint, jsonify, request, send_file

from app.models import (
    Grade,
    Homework,
    ScheduleEntry,
    ScheduleTeacherLink,
    StudentProfile,
    TeacherProfile,
    User,
)
from app.services.auth_service import decode_jwt
from app.utils.student_identity import extract_student_code, resolve_student_user


exports_bp = Blueprint("exports", __name__)
ALLOWED_FORMATS = {"html", "pdf", "xlsx"}
ALLOWED_LANGUAGES = {"ru", "en", "kg"}
DAY_ORDER = {
    "monday": 1,
    "tuesday": 2,
    "wednesday": 3,
    "thursday": 4,
    "friday": 5,
    "saturday": 6,
    "понедельник": 1,
    "вторник": 2,
    "среда": 3,
    "четверг": 4,
    "пятница": 5,
    "суббота": 6,
    "дүйшөмбү": 1,
    "шейшемби": 2,
    "шаршемби": 3,
    "бейшемби": 4,
    "жума": 5,
    "ишемби": 6,
}
DAY_CANONICAL = {
    "monday": "monday",
    "понедельник": "monday",
    "дүйшөмбү": "monday",
    "tuesday": "tuesday",
    "вторник": "tuesday",
    "шейшемби": "tuesday",
    "wednesday": "wednesday",
    "среда": "wednesday",
    "шаршемби": "wednesday",
    "thursday": "thursday",
    "четверг": "thursday",
    "бейшемби": "thursday",
    "friday": "friday",
    "пятница": "friday",
    "жума": "friday",
    "saturday": "saturday",
    "суббота": "saturday",
    "ишемби": "saturday",
}
DAY_TRANSLATIONS = {
    "ru": {
        "monday": "Понедельник",
        "tuesday": "Вторник",
        "wednesday": "Среда",
        "thursday": "Четверг",
        "friday": "Пятница",
        "saturday": "Суббота",
    },
    "en": {
        "monday": "Monday",
        "tuesday": "Tuesday",
        "wednesday": "Wednesday",
        "thursday": "Thursday",
        "friday": "Friday",
        "saturday": "Saturday",
    },
    "kg": {
        "monday": "Дүйшөмбү",
        "tuesday": "Шейшемби",
        "wednesday": "Шаршемби",
        "thursday": "Бейшемби",
        "friday": "Жума",
        "saturday": "Ишемби",
    },
}
EXPORT_TEXT = {
    "ru": {
        "no_data": "Нет данных",
        "generated": "Сформировано",
        "student_id": "ID студента",
        "group": "Группа",
        "grades_title": "Отчёт по оценкам",
        "schedule_title": "Отчёт по расписанию",
        "homework_title": "Домашнее задание",
        "general_subject": "Общее",
        "col_date": "Дата",
        "col_subject": "Предмет",
        "col_grade": "Оценка",
        "col_teacher": "Преподаватель",
        "col_day": "День",
        "col_start": "Начало",
        "col_end": "Конец",
        "col_room": "Аудитория",
        "col_group": "Группа",
        "col_title": "Заголовок",
        "col_description": "Описание",
        "col_due_date": "Срок сдачи",
        "col_created_at": "Создано",
    },
    "en": {
        "no_data": "No data",
        "generated": "Generated",
        "student_id": "Student ID",
        "group": "Group",
        "grades_title": "Grades report",
        "schedule_title": "Schedule report",
        "homework_title": "Homework",
        "general_subject": "General",
        "col_date": "Date",
        "col_subject": "Subject",
        "col_grade": "Grade",
        "col_teacher": "Teacher",
        "col_day": "Day",
        "col_start": "Start",
        "col_end": "End",
        "col_room": "Room",
        "col_group": "Group",
        "col_title": "Title",
        "col_description": "Description",
        "col_due_date": "Due date",
        "col_created_at": "Created at",
    },
    "kg": {
        "no_data": "Маалымат жок",
        "generated": "Түзүлгөн",
        "student_id": "Студент ID",
        "group": "Топ",
        "grades_title": "Баалар боюнча отчет",
        "schedule_title": "Жадыбал боюнча отчет",
        "homework_title": "Үй тапшырма",
        "general_subject": "Жалпы",
        "col_date": "Дата",
        "col_subject": "Сабак",
        "col_grade": "Баа",
        "col_teacher": "Мугалим",
        "col_day": "Күн",
        "col_start": "Башталышы",
        "col_end": "Аягы",
        "col_room": "Аудитория",
        "col_group": "Топ",
        "col_title": "Аталышы",
        "col_description": "Сүрөттөмө",
        "col_due_date": "Тапшыруу мөөнөтү",
        "col_created_at": "Түзүлгөн",
    },
}


def _get_format():
    raw = str(request.args.get("format") or "pdf").strip().lower()
    if raw == "excel":
        raw = "xlsx"
    return raw


def _get_lang():
    raw = str(request.args.get("lang") or "").strip().lower()
    if raw in ALLOWED_LANGUAGES:
        return raw

    accept = str(request.headers.get("Accept-Language") or "").strip().lower()
    if accept:
        token = accept.split(",")[0].split("-")[0].strip()
        if token in ALLOWED_LANGUAGES:
            return token

    return "ru"


def _tr(lang, key):
    table = EXPORT_TEXT.get(lang) or EXPORT_TEXT["ru"]
    return table.get(key) or EXPORT_TEXT["ru"].get(key) or key


def _localize_day(day_of_week, lang):
    source = str(day_of_week or "").strip()
    canonical = DAY_CANONICAL.get(source.lower())
    if not canonical:
        return source
    return DAY_TRANSLATIONS.get(lang, DAY_TRANSLATIONS["ru"]).get(canonical, source)


def _error_response(message, code):
    return jsonify({"error": message}), code


def _extract_bearer_token():
    auth_header = request.headers.get("Authorization", "")
    if auth_header.startswith("Bearer "):
        token = auth_header.split(" ", 1)[1].strip()
        if token:
            return token
    return None


def _resolve_current_user():
    token = _extract_bearer_token() or str(request.args.get("token") or "").strip()
    if not token:
        return None, _error_response("Authorization token is missing", 401)

    try:
        payload = decode_jwt(token)
        user_id = int(payload.get("sub"))
        user = User.query.get(user_id)
        if not user:
            return None, _error_response("User not found", 401)
        return user, None
    except jwt.ExpiredSignatureError:
        return None, _error_response("Token expired", 401)
    except jwt.InvalidTokenError:
        return None, _error_response("Invalid token", 401)
    except (TypeError, ValueError):
        return None, _error_response("Invalid token payload", 401)


def _profile_name(user: User, teacher_profiles, student_profiles):
    if user.role == "teacher":
        profile = teacher_profiles.get(user.id)
    elif user.role == "student":
        profile = student_profiles.get(user.id)
    else:
        profile = None

    if profile:
        full_name = " ".join(
            [part for part in [profile.last_name, profile.first_name, profile.middle_name] if part]
        ).strip()
        if full_name:
            return full_name
    return user.login


def _build_html(title, subtitle, columns, rows, no_data_text, lang):
    table_header = "".join([f"<th>{html.escape(col)}</th>" for col in columns])

    if rows:
        table_rows = []
        for row in rows:
            cells = "".join([f"<td>{html.escape(str(value))}</td>" for value in row])
            table_rows.append(f"<tr>{cells}</tr>")
        table_body = "".join(table_rows)
    else:
        table_body = f'<tr><td colspan="{len(columns)}">{html.escape(no_data_text)}</td></tr>'

    return f"""<!doctype html>
<html lang="{html.escape(lang)}">
<head>
  <meta charset="utf-8" />
  <meta name="color-scheme" content="only light" />
  <title>{html.escape(title)}</title>
  <style>
    :root {{
      color-scheme: only light;
    }}
    body {{
      font-family: "Times New Roman", serif;
      color: #111 !important;
      margin: 24px;
      background: #fff !important;
    }}
    h1 {{ margin: 0 0 6px 0; font-size: 24px; }}
    p {{ margin: 0 0 18px 0; color: #333; font-size: 13px; }}
    table {{
      width: 100%;
      border-collapse: collapse;
      table-layout: fixed;
    }}
    th, td {{
      border: 1px solid #111 !important;
      padding: 8px;
      font-size: 13px;
      vertical-align: top;
      word-wrap: break-word;
      color: #111 !important;
      background: #fff !important;
    }}
    th {{ background: #f1f1f1 !important; text-align: left; }}
    @media (prefers-color-scheme: dark) {{
      html, body, table, thead, tbody, tr, th, td, p, h1 {{
        color: #111 !important;
        background: #fff !important;
        border-color: #111 !important;
      }}
      th {{
        background: #f1f1f1 !important;
      }}
    }}
  </style>
</head>
<body>
  <h1>{html.escape(title)}</h1>
  <p>{html.escape(subtitle)}</p>
  <table>
    <thead>
      <tr>{table_header}</tr>
    </thead>
    <tbody>
      {table_body}
    </tbody>
  </table>
</body>
</html>"""


def _build_excel(title, columns, rows):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Export"

    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
    ws["A1"].font = Font(name="Times New Roman", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.append(columns)
    header_row = 2
    header_fill = PatternFill("solid", fgColor="EDEDED")
    for idx in range(1, len(columns) + 1):
        cell = ws.cell(row=header_row, column=idx)
        cell.font = Font(name="Times New Roman", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in rows:
        ws.append(row)

    for idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row_index in range(3, 3 + len(rows)):
            value = ws.cell(row=row_index, column=idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
            ws.cell(row=row_index, column=idx).alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
        ws.column_dimensions[chr(64 + idx)].width = min(max(max_len + 2, 12), 45)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _resolve_pdf_font_name():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "Helvetica"
    candidates = [
        ("DejaVuSans", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        ("DejaVuSans", "/usr/share/fonts/dejavu/DejaVuSans.ttf"),
        ("NotoSans", "/usr/share/fonts/google-noto-vf/NotoSans[wght].ttf"),
    ]
    for candidate_name, path in candidates:
        if not os.path.exists(path):
            continue
        font_name = candidate_name
        if font_name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(TTFont(font_name, path))
        break

    return font_name


def _build_pdf(title, subtitle, columns, rows, no_data_text):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    font_name = _resolve_pdf_font_name()
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Title"],
        fontName=font_name,
        textColor=colors.black,
    )
    normal_style = ParagraphStyle(
        "ExportNormal",
        parent=styles["Normal"],
        fontName=font_name,
        textColor=colors.black,
    )
    content = [
        Paragraph(f"<b>{html.escape(title)}</b>", title_style),
        Spacer(1, 4),
        Paragraph(html.escape(subtitle), normal_style),
        Spacer(1, 10),
    ]

    table_data = [columns]
    table_data.extend(rows if rows else [[no_data_text] + [""] * (len(columns) - 1)])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("ALIGN", (0, 0), (-1, 0), "LEFT"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    content.append(table)
    doc.build(content)
    return buffer.getvalue()


def _build_file_bytes(export_format, title, subtitle, columns, rows, no_data_text, lang):
    if export_format == "html":
        return _build_html(title, subtitle, columns, rows, no_data_text, lang).encode("utf-8")
    if export_format == "xlsx":
        return _build_excel(title, columns, rows)
    return _build_pdf(title, subtitle, columns, rows, no_data_text)


def _file_meta(export_format):
    if export_format == "html":
        return "text/html; charset=utf-8", "html"
    if export_format == "xlsx":
        return (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "xlsx",
        )
    return "application/pdf", "pdf"


def _schedule_day_rank(day_of_week):
    value = str(day_of_week or "").strip().lower()
    return DAY_ORDER.get(value, 99)


@exports_bp.get("/exports/grades/<student_ref>")
def export_grades(student_ref: str):
    user, auth_error = _resolve_current_user()
    if auth_error:
        return auth_error
    lang = _get_lang()

    if user.role not in {"admin", "teacher", "student"}:
        return _error_response("Forbidden", 403)

    preferred_group = user.group_id if user.role == "student" else None
    student, resolve_error = resolve_student_user(student_ref, preferred_group=preferred_group)
    if resolve_error == "ambiguous":
        return _error_response("Student identifier is ambiguous. Use full login", 409)
    if not student:
        return _error_response("Student not found", 404)
    if user.role == "student" and user.id != student.id:
        return _error_response("Forbidden", 403)

    export_format = _get_format()
    if export_format not in ALLOWED_FORMATS:
        return _error_response("Unsupported format. Use html, pdf, xlsx", 400)

    rows = Grade.query.filter_by(student_id=student.id).order_by(Grade.created_at.desc()).all()
    teacher_ids = list({row.teacher_id for row in rows})
    teachers = User.query.filter(User.id.in_(teacher_ids)).all() if teacher_ids else []
    teacher_profiles = TeacherProfile.query.filter(TeacherProfile.user_id.in_(teacher_ids)).all() if teacher_ids else []

    teacher_map = {row.id: row for row in teachers}
    teacher_profile_map = {row.user_id: row for row in teacher_profiles}
    student_profiles = StudentProfile.query.filter_by(user_id=student.id).all()
    student_profile_map = {row.user_id: row for row in student_profiles}

    table_rows = []
    for row in rows:
        teacher = teacher_map.get(row.teacher_id)
        teacher_name = teacher.login if teacher else "-"
        if teacher:
            teacher_name = _profile_name(teacher, teacher_profile_map, {})
        table_rows.append(
            [
                row.created_at.strftime("%Y-%m-%d %H:%M"),
                row.subject,
                row.value,
                teacher_name,
            ]
        )

    student_name = _profile_name(student, {}, student_profile_map)
    student_code = extract_student_code(student.login) or str(student.id)
    title = f"{_tr(lang, 'grades_title')} - {student_name}"
    subtitle = (
        f"{_tr(lang, 'student_id')}: {student_code} | "
        f"{_tr(lang, 'generated')}: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    columns = [
        _tr(lang, "col_date"),
        _tr(lang, "col_subject"),
        _tr(lang, "col_grade"),
        _tr(lang, "col_teacher"),
    ]

    content = _build_file_bytes(
        export_format,
        title,
        subtitle,
        columns,
        table_rows,
        _tr(lang, "no_data"),
        lang,
    )
    mime_type, extension = _file_meta(export_format)
    file_name = f"grades_student_{student_code}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{extension}"

    return send_file(
        BytesIO(content),
        mimetype=mime_type,
        as_attachment=True,
        download_name=file_name,
    )


@exports_bp.get("/exports/schedule/<group_id>")
def export_schedule(group_id: str):
    user, auth_error = _resolve_current_user()
    if auth_error:
        return auth_error
    lang = _get_lang()

    if user.role == "student":
        if not user.group_id or str(user.group_id) != str(group_id):
            return _error_response("Forbidden", 403)
    elif user.role not in {"admin", "teacher", "scheduler"}:
        return _error_response("Forbidden", 403)

    export_format = _get_format()
    if export_format not in ALLOWED_FORMATS:
        return _error_response("Unsupported format. Use html, pdf, xlsx", 400)

    rows = ScheduleEntry.query.filter_by(group_id=str(group_id)).all()
    rows = sorted(rows, key=lambda item: (_schedule_day_rank(item.day_of_week), item.start_time))

    schedule_ids = [row.id for row in rows]
    links = (
        ScheduleTeacherLink.query.filter(ScheduleTeacherLink.schedule_entry_id.in_(schedule_ids)).all()
        if schedule_ids
        else []
    )
    teacher_ids = list({row.teacher_id for row in links})
    teachers = User.query.filter(User.id.in_(teacher_ids)).all() if teacher_ids else []
    teacher_profiles = TeacherProfile.query.filter(TeacherProfile.user_id.in_(teacher_ids)).all() if teacher_ids else []

    link_map = {row.schedule_entry_id: row for row in links}
    teacher_map = {row.id: row for row in teachers}
    teacher_profile_map = {row.user_id: row for row in teacher_profiles}

    table_rows = []
    for row in rows:
        link = link_map.get(row.id)
        teacher_name = "-"
        if link:
            teacher = teacher_map.get(link.teacher_id)
            if teacher:
                teacher_name = _profile_name(teacher, teacher_profile_map, {})

        table_rows.append(
            [
                _localize_day(row.day_of_week, lang),
                row.start_time,
                row.end_time,
                row.subject,
                teacher_name,
                row.room or "-",
            ]
        )

    title = f"{_tr(lang, 'schedule_title')} - {_tr(lang, 'group')} {group_id}"
    subtitle = f"{_tr(lang, 'generated')}: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    columns = [
        _tr(lang, "col_day"),
        _tr(lang, "col_start"),
        _tr(lang, "col_end"),
        _tr(lang, "col_subject"),
        _tr(lang, "col_teacher"),
        _tr(lang, "col_room"),
    ]

    content = _build_file_bytes(
        export_format,
        title,
        subtitle,
        columns,
        table_rows,
        _tr(lang, "no_data"),
        lang,
    )
    mime_type, extension = _file_meta(export_format)
    file_name = f"schedule_{group_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{extension}"

    return send_file(
        BytesIO(content),
        mimetype=mime_type,
        as_attachment=True,
        download_name=file_name,
    )


@exports_bp.get("/exports/homework/<int:homework_id>")
def export_homework(homework_id: int):
    user, auth_error = _resolve_current_user()
    if auth_error:
        return auth_error
    lang = _get_lang()

    if user.role not in {"admin", "teacher", "student"}:
        return _error_response("Forbidden", 403)

    row = Homework.query.get(homework_id)
    if not row:
        return _error_response("Homework not found", 404)

    if user.role == "student" and str(user.group_id or "") != str(row.group_id):
        return _error_response("Forbidden", 403)

    export_format = _get_format()
    if export_format not in ALLOWED_FORMATS:
        return _error_response("Unsupported format. Use html, pdf, xlsx", 400)

    teacher = User.query.get(row.teacher_id)
    teacher_name = teacher.login if teacher else "-"

    if teacher:
        teacher_profile = TeacherProfile.query.filter_by(user_id=teacher.id).first()
        teacher_profile_map = {teacher.id: teacher_profile} if teacher_profile else {}
        teacher_name = _profile_name(teacher, teacher_profile_map, {})

    table_rows = [
        [
            row.group_id,
            row.subject or _tr(lang, "general_subject"),
            row.title,
            row.description,
            row.due_date or "-",
            teacher_name,
            row.created_at.strftime("%Y-%m-%d %H:%M"),
        ]
    ]

    title = f"{_tr(lang, 'homework_title')} - {row.subject or _tr(lang, 'general_subject')}"
    subtitle = (
        f"{_tr(lang, 'group')}: {row.group_id} | "
        f"{_tr(lang, 'generated')}: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    columns = [
        _tr(lang, "col_group"),
        _tr(lang, "col_subject"),
        _tr(lang, "col_title"),
        _tr(lang, "col_description"),
        _tr(lang, "col_due_date"),
        _tr(lang, "col_teacher"),
        _tr(lang, "col_created_at"),
    ]

    content = _build_file_bytes(
        export_format,
        title,
        subtitle,
        columns,
        table_rows,
        _tr(lang, "no_data"),
        lang,
    )
    mime_type, extension = _file_meta(export_format)
    file_name = f"homework_{row.id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{extension}"

    return send_file(
        BytesIO(content),
        mimetype=mime_type,
        as_attachment=True,
        download_name=file_name,
    )
